"""
Парсер Excel файлов для загрузки прайс-листов
"""
import io
import hashlib
import tempfile
import os
from typing import List, Dict, Optional, Any
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image
from ...models.product import ProductCreate, Product
from ...services.image_service import image_service
from ...services.product_service import ProductService


class ExcelParser:
    """
    Парсер Excel файлов с прайс-листами продуктов.

    Класс используется в API загрузки прайс-листа как
    `ExcelParser(content, settings.UPLOAD_DIR)` с последующим вызовом
    `parser.parse(...)`, поэтому поддерживает как старый, так и новый интерфейсы.
    """

    COLUMN_MAPPING = {
        "name": ["Наименование", "наименование", "Название", "название"],
        "image": ["Изображение устройства", "изображение устройства", "Изображение", "изображение"],
        "description": ["Характеристика", "характеристика", "Описание", "описание"],
        "price_uzs": ["Наличные", "наличные", "Цена UZS", "цена uzs"],
        "price_usd": ["USD", "usd", "Цена USD", "цена usd"],
        "usd_rate": ["курс", "Курс", "Курс USD", "курс usd"]
    }

    # Ключевые слова для определения типа товара по названию/описанию
    CATEGORY_KEYWORDS = {
        "nvr": [
            "nvr",
            "сетевой видеорегистратор",
            "ip-видеорегистратор",
            "ip видеорегистратор",
        ],
        "dvr": [
            "dvr",
            "цифровой видеорегистратор",
            "гибридный видеорегистратор",
        ],
        "ip_cameras": [
            "ip-камера",
            "ip камера",
            "ip-камеры",
            "ip видеокамера",
            "ip видеокамеры",
            "сетевой камера",
            "сетевые камеры",
        ],
        "analog": [
            "аналоговая камера",
            "аналоговые камеры",
            "turbo hd",
            "hd-tvi",
            "ahd",
            "cvi",
        ],
        "accessories": [
            "жесткий диск",
            "жёсткий диск",
            "hdd",
            "ssd",
            "кронштейн",
            "блок питания",
            "источник питания",
            "кабель",
            "разъём",
            "разьем",
            "адаптер",
            "микрофон",
            "объектив",
        ],
    }

    def __init__(self, content: bytes | None = None, upload_dir: str | None = None):
        """
        Args:
            content: Сырые байты Excel-файла (новый API).
            upload_dir: Каталог для временного сохранения файла.

        Если `content` не указан, класс можно использовать по-старому,
        передавая путь к файлу в методы `parse_excel` / `parse_and_create_products`.
        """
        self.column_mapping: Dict[str, int] = {}
        self.product_service = ProductService()
        # Последний считанный курс USD из файла (нужен, если курс указан только в одной строке)
        self._last_usd_rate: Optional[float] = None
        self._temp_file_path: Optional[str] = None
        # Кэш изображений текущего листа: строка Excel -> список изображений
        self._image_ws = None
        self._image_map: Dict[int, List[Dict[str, Any]]] = {}
        self._claimed_images: set[int] = set()
        # Кэш загрузок: md5 содержимого -> публичный URL (не грузим повторы дважды)
        self._uploaded_by_hash: Dict[str, str] = {}

        if content is not None:
            # Новый сценарий: сохраняем контент во временный файл
            if upload_dir is None:
                upload_dir = "data/uploads"

            os.makedirs(upload_dir, exist_ok=True)

            fd, path = tempfile.mkstemp(suffix=".xlsx", dir=upload_dir)
            with os.fdopen(fd, "wb") as f:
                f.write(content)

            self._temp_file_path = path

    def parse(
        self,
        sheet_name: Optional[str] = None,
        sheet_index: int = 0,
        sheet_names: Optional[List[str]] = None,
        skip_duplicates: bool = True,
    ) -> Dict[str, Any]:
        """
        Новый API, который ожидает FastAPI-роут загрузки прайс-листа.

        Параметры `sheet_name`, `sheet_index`, `skip_duplicates` пока
        не используются, но оставлены для совместимости.
        """
        if not self._temp_file_path:
            raise ValueError(
                "Временный файл не инициализирован. Передайте content в конструктор ExcelParser."
            )

        errors: List[str] = []
        products_dicts: List[Dict[str, Any]] = []

        try:
            # data_only=True: для формул (=F4*G4, =NVR!G4) берём вычисленное
            # значение из кэша Excel, а не текст формулы
            workbook = load_workbook(self._temp_file_path, data_only=True)

            # Определяем, какие листы нужно парсить
            if sheet_names:
                target_sheets = [
                    workbook[s]
                    for s in sheet_names
                    if s in workbook.sheetnames
                ]
                missing = [s for s in sheet_names if s not in workbook.sheetnames]
                for name in missing:
                    errors.append(f"Лист '{name}' не найден в файле")
            elif sheet_name and sheet_name in workbook.sheetnames:
                target_sheets = [workbook[sheet_name]]
            elif 0 <= sheet_index < len(workbook.sheetnames):
                target_sheets = [workbook.worksheets[sheet_index]]
            else:
                target_sheets = [workbook.active]

            processed_rows = 0
            skipped_rows = 0
            duplicate_rows = 0
            total_rows = 0

            seen_names: set[str] = set()

            for ws in target_sheets:
                # Для каждого листа ищем свои колонки
                if not self._find_columns(ws):
                    errors.append(
                        f"Не удалось найти необходимые колонки на листе '{ws.title}'"
                    )
                    continue

                # Сбрасываем курс USD между листами
                self._last_usd_rate = None

                for row_num in range(2, ws.max_row + 1):
                    total_rows += 1
                    try:
                        product = self._parse_row(ws, row_num)
                        if not product:
                            skipped_rows += 1
                            continue

                        name_key = product.name.strip()
                        if skip_duplicates and name_key in seen_names:
                            duplicate_rows += 1
                            continue

                        seen_names.add(name_key)
                        products_dicts.append(product.dict())
                        processed_rows += 1
                    except Exception as e:
                        errors.append(
                            f"Ошибка в листе '{ws.title}', строка {row_num}: {str(e)}"
                        )

            stats = {
                "total_rows": total_rows,
                "processed": processed_rows,
                "created": 0,  # Перезапишется на этапе сохранения в БД
                "skipped": skipped_rows,
                "duplicates": duplicate_rows,
                "errors": len(errors),
            }

            return {
                "success": True,
                "products": products_dicts,
                "stats": stats,
                "errors": errors,
                "sheet_name": sheet_name or (target_sheets[0].title if target_sheets else None),
                "sheet_names": [ws.title for ws in target_sheets],
            }

        except Exception as e:
            return {
                "success": False,
                "products": [],
                "stats": {
                    "total_rows": 0,
                    "processed": 0,
                    "created": 0,
                    "skipped": 0,
                    "duplicates": 0,
                    "errors": 1,
                },
                "errors": [str(e)],
                "sheet_name": sheet_name,
                "sheet_names": sheet_names or [],
            }

    def parse_excel(self, file_path: str) -> List[ProductCreate]:
        """
        Парсит Excel файл и возвращает список продуктов

        Args:
            file_path: Путь к Excel файлу

        Returns:
            Список объектов ProductCreate
        """
        try:
            workbook = load_workbook(file_path, data_only=True)
            worksheet = workbook.active

            if not self._find_columns(worksheet):
                raise ValueError("Не удалось найти необходимые колонки в файле")

            products = []

            for row_num in range(2, worksheet.max_row + 1):
                try:
                    product_data = self._parse_row(worksheet, row_num)
                    if product_data:
                        products.append(product_data)
                except Exception as e:
                    print(f"[WARNING] Пропуск строки {row_num}: {str(e)}")
                    continue

            return products

        except Exception as e:
            print(f"[ERROR] Ошибка парсинга Excel файла: {str(e)}")
            raise

    def _find_columns(self, worksheet) -> bool:
        """
        Находит колонки по заголовкам с гибким сопоставлением.

        В реальных прайс-листах часто есть «шапка» с логотипом и описанием
        над строкой заголовков, поэтому мы ищем подходящую строку среди
        первых нескольких рядов, а не только в самом первом.
        """
        max_header_row = min(10, worksheet.max_row)
        self.column_mapping.clear()

        for row_idx in range(1, max_header_row + 1):
            row = worksheet[row_idx]

            # Временное сопоставление для этой строки
            temp_mapping: Dict[str, int] = {}

            for col_idx, cell in enumerate(row, 1):
                if cell.value:
                    header_value = str(cell.value).strip()

                    for field_name, possible_names in self.COLUMN_MAPPING.items():
                        if header_value in possible_names:
                            temp_mapping[field_name] = col_idx
                            break

            # Считаем строку заголовков подходящей, если в ней найдено
            # хотя бы 3 ключевых колонки (например, name, price_uzs, description и т.п.)
            if len(temp_mapping) >= 3:
                self.column_mapping = temp_mapping
                break

        missing_columns = set(self.COLUMN_MAPPING.keys()) - set(self.column_mapping.keys())
        if missing_columns:
            print(f"[ERROR] Отсутствуют колонки: {missing_columns}")
            return False
        return True


    def _parse_row(self, worksheet, row_num: int) -> Optional[ProductCreate]:
        """Парсит одну строку и возвращает ProductCreate.
        
        Условие: парсим все товары, у которых есть название и хотя бы одна цена.
        - «Наличные» -> price_uzs (сум);
        - «USD» -> price_usd (доллары);
        - курс usd_rate берём из столбца «Курс», а если он пуст и заданы обе цены,
        считаем как price_uzs / price_usd.
        """
        name = self._get_cell_value(worksheet, row_num, "name")
        if not name:
            return None

        description = self._get_cell_value(worksheet, row_num, "description") or ""
        raw_price_uzs = self._get_numeric_value(worksheet, row_num, "price_uzs")
        raw_price_usd = self._get_numeric_value(worksheet, row_num, "price_usd")
        raw_usd_rate = self._get_numeric_value(worksheet, row_num, "usd_rate")

        # Если курс не указан в этой строке, но уже встречался выше — используем последний.
        if raw_usd_rate is None and self._last_usd_rate is not None:
            raw_usd_rate = self._last_usd_rate
        # Запоминаем курс, если он есть в текущей строке.
        if raw_usd_rate is not None:
            self._last_usd_rate = raw_usd_rate

        # Должна быть хотя бы одна цена
        if raw_price_uzs is None and raw_price_usd is None:
            return None

        # Если есть только одна цена - вычисляем вторую через курс
        if raw_price_uzs is not None and raw_price_usd is None:
            price_uzs = float(raw_price_uzs)
            if raw_usd_rate and raw_usd_rate > 0:
                price_usd = price_uzs / raw_usd_rate
                usd_rate = raw_usd_rate
            else:
                price_usd = price_uzs  # Если нет курса, дублируем цену
                usd_rate = 1.0
        elif raw_price_usd is not None and raw_price_uzs is None:
            price_usd = float(raw_price_usd)
            if raw_usd_rate and raw_usd_rate > 0:
                price_uzs = price_usd * raw_usd_rate
                usd_rate = raw_usd_rate
            else:
                price_uzs = price_usd  # Если нет курса, дублируем цену
                usd_rate = 1.0
        else:
            # Обе цены указаны
            price_uzs = float(raw_price_uzs)
            price_usd = float(raw_price_usd)
            
            # Вычисляем курс
            if raw_usd_rate is not None:
                usd_rate = float(raw_usd_rate)
            elif price_uzs > 0 and price_usd > 0:
                usd_rate = price_uzs / price_usd
            else:
                usd_rate = 1.0

        image_url = self._extract_and_upload_image(worksheet, row_num)

        # Определяем категорию по названию и описанию
        category = self._infer_category(str(name), str(description))

        return ProductCreate(
            name=str(name).strip(),
            description=str(description).strip(),
            price_uzs=price_uzs,
            price_usd=price_usd,
            usd_rate=usd_rate,
            image_url=image_url,
            category=category,
        )

    def _infer_category(self, name: str, description: str) -> str:
        """
        Определяет категорию товара по ключевым словам в названии и описании.

        Возможные значения:
            - nvr
            - dvr
            - ip_cameras
            - analog
            - accessories
        """
        text = f"{name} {description}".lower()

        for category, keywords in self.CATEGORY_KEYWORDS.items():
            for kw in keywords:
                if kw in text:
                    return category

        # Если ничего не нашли — считаем, что это аксессуар,
        # чтобы всегда возвращать одну из разрешённых категорий.
        return "accessories"

    def _get_cell_value(self, worksheet, row_num: int, field_name: str) -> Any:
        """Получает значение ячейки по имени поля"""
        col_num = self.column_mapping.get(field_name)
        if col_num:
            cell = worksheet.cell(row=row_num, column=col_num)
            return cell.value
        return None

    def _get_numeric_value(self, worksheet, row_num: int, field_name: str) -> Optional[float]:
        """Получает числовое значение из ячейки"""
        value = self._get_cell_value(worksheet, row_num, field_name)
        if value is not None:
            try:
                # Поддержка строковых значений с пробелами и запятыми как разделителем дробной части
                if isinstance(value, str):
                    cleaned = (
                        value.replace("\xa0", " ")  # неразрывные пробелы
                        .replace(" ", "")  # убираем разделители тысяч
                        .replace(",", ".")  # локальные десятичные запятые
                    )
                    return float(cleaned)
                return float(value)
            except (ValueError, TypeError):
                return None
        return None

    def _extract_and_upload_image(self, worksheet, row_num: int) -> Optional[str]:
        """
        Извлекает изображение, привязанное к строке Excel, и загружает его
        в Supabase Storage через `image_service`.

        Возвращает публичный URL или None, если картинка отсутствует
        или произошла ошибка при загрузке.
        """
        try:
            # Карту изображений строим один раз на лист, а не на каждую строку
            if self._image_ws is not worksheet:
                self._image_ws = worksheet
                self._image_map = self._build_image_map(worksheet)
                self._claimed_images = set()

            candidates = list(self._image_map.get(row_num, []))

            # Fallback на соседние строки — но только если соседняя строка
            # сама не является товаром, иначе мы заберём её картинку.
            if not candidates:
                for neighbor in (row_num - 1, row_num + 1):
                    if neighbor < 2:
                        continue
                    if self._get_cell_value(worksheet, neighbor, "name"):
                        continue
                    candidates.extend(self._image_map.get(neighbor, []))

            candidates = [
                c for c in candidates if id(c["image"]) not in self._claimed_images
            ]

            # Если колонка изображений известна, предпочитаем картинки из неё
            img_col = self.column_mapping.get("image")
            if img_col:
                in_column = [c for c in candidates if abs(c["col"] - img_col) <= 1]
                if in_column:
                    candidates = in_column

            for entry in candidates:
                # Используем отдельную папку, чтобы различать импорт из Excel.
                url = self._upload_image_to_storage(entry["image"], folder="excel-import")
                if url:
                    self._claimed_images.add(id(entry["image"]))
                    return url

            return None

        except Exception as e:
            print(f"[WARNING] Ошибка обработки изображения в строке {row_num}: {str(e)}")
            return None

    def _build_image_map(self, worksheet) -> Dict[int, List[Dict[str, Any]]]:
        """
        Строит отображение «строка Excel (1-индексная) -> изображения».

        Якоря openpyxl 0-индексные, поэтому +1. Строку определяем по центру
        картинки (учитывая rowOff и нижний угол twoCellAnchor): верхний угол
        часто «залезает» в строку выше, и по нему одному строка определяется
        неверно.
        """
        image_map: Dict[int, List[Dict[str, Any]]] = {}

        for image in worksheet._images:
            if not isinstance(image, OpenpyxlImage):
                continue

            anchor = getattr(image, "anchor", None)
            frm = getattr(anchor, "_from", None)
            if frm is None:
                # AbsoluteAnchor не привязан к ячейке — сопоставить со строкой нельзя
                print("[WARNING] Изображение с абсолютным якорем пропущено")
                continue

            pos_from = frm.row + self._row_offset_ratio(worksheet, frm)
            to = getattr(anchor, "to", None)
            if to is not None:
                pos_to = to.row + self._row_offset_ratio(worksheet, to)
                center = (pos_from + pos_to) / 2
            else:
                # OneCellAnchor: считаем, что картинка занимает примерно одну строку
                center = pos_from + 0.5

            excel_row = int(center) + 1
            image_map.setdefault(excel_row, []).append({
                "image": image,
                "col": frm.col + 1,
            })

        return image_map

    def _row_offset_ratio(self, worksheet, marker) -> float:
        """Доля строки, на которую смещён якорь (rowOff в EMU / высота строки)."""
        dim = worksheet.row_dimensions.get(marker.row + 1)
        height_pt = dim.height if dim is not None and dim.height else None
        if not height_pt:
            height_pt = worksheet.sheet_format.defaultRowHeight or 15.0
        height_emu = float(height_pt) * 12700  # 1 pt = 12700 EMU
        return min(marker.rowOff / height_emu, 1.0) if height_emu else 0.0

    def _upload_image_to_storage(
        self,
        openpyxl_image: OpenpyxlImage,
        folder: str = "products",
    ) -> Optional[str]:
        """
        Извлекает байты изображения из объекта openpyxl и загружает их
        в Supabase Storage через `image_service`.

        Args:
            openpyxl_image: Объект изображения из openpyxl.
            folder: Логическая папка внутри bucket (например, `products` или
                    `excel-import`), прокидывается в Supabase Storage.

        Returns:
            Публичный URL загруженного файла или None в случае ошибки.
        """
        try:
            # В openpyxl сами байты изображения лежат во внутреннем объекте `_data()`,
            # а `ref` / `anchor` относятся к позиционированию на листе.
            raw_data = openpyxl_image._data()

            # В разных версиях openpyxl `_data()` может возвращать объект с полем `blob`
            # или сразу bytes. Нормализуем к bytes.
            if hasattr(raw_data, "blob"):
                image_bytes_raw = raw_data.blob
            elif isinstance(raw_data, (bytes, bytearray)):
                image_bytes_raw = bytes(raw_data)
            else:
                # На всякий случай пробуем получить атрибут `fp` (файловый объект)
                fp = getattr(raw_data, "fp", None)
                if fp and hasattr(fp, "read"):
                    image_bytes_raw = fp.read()
                else:
                    raise ValueError("Неподдерживаемый формат данных изображения из openpyxl")

            # Повторяющиеся картинки (один товар в нескольких строках) грузим один раз
            content_hash = hashlib.md5(image_bytes_raw).hexdigest()
            if content_hash in self._uploaded_by_hash:
                return self._uploaded_by_hash[content_hash]

            pil_image = Image.open(io.BytesIO(image_bytes_raw))
            if pil_image.mode not in ("RGB", "RGBA", "L", "P"):
                # PNG не поддерживает, например, CMYK
                pil_image = pil_image.convert("RGB")

            buffer = io.BytesIO()
            pil_image.save(buffer, "PNG")

            # `image_service` — это обёртка над Supabase Storage (см. ImageKitService),
            # здесь мы просто передаём байты и путь для загрузки.
            url = image_service.upload_image(
                buffer.getvalue(), f"product_{content_hash}.png", folder=folder
            )
            if url:
                self._uploaded_by_hash[content_hash] = url
            return url

        except Exception as e:
            print(f"[ERROR] Ошибка загрузки изображения в Supabase Storage: {str(e)}")
            return None

    async def parse_and_create_products(self, file_path: str) -> Dict[str, Any]:
        """
        Парсит Excel файл и создает продукты в базе данных

        Args:
            file_path: Путь к Excel файлу

        Returns:
            Словарь со статистикой создания
        """
        try:
            print("[INFO] Начало парсинга Excel файла...")
            products_data = self.parse_excel(file_path)

            if not products_data:
                return {
                    "success": False,
                    "message": "Не найдены валидные продукты для создания",
                    "total_parsed": 0,
                    "created_count": 0,
                    "failed_count": 0,
                    "products": []
                }

            print(f"[INFO] Распарсено {len(products_data)} продуктов, создание в базе данных...")

            created_products = await self.product_service.create_products_bulk(products_data)

            result = {
                "success": True,
                "message": f"Успешно создано {len(created_products)} продуктов",
                "total_parsed": len(products_data),
                "created_count": len(created_products),
                "failed_count": len(products_data) - len(created_products),
                "products": [
                    {
                        "id": product.id,
                        "name": product.name,
                        "price_usd": product.price_usd,
                        "image_url": product.image_url
                    } for product in created_products
                ]
            }

            print(f"[SUCCESS] Создано {len(created_products)} продуктов в базе данных")
            return result

        except Exception as e:
            error_message = f"Ошибка создания продуктов: {str(e)}"
            print(f"[ERROR] {error_message}")
            return {
                "success": False,
                "message": error_message,
                "total_parsed": 0,
                "created_count": 0,
                "failed_count": 0,
                "products": []
            }
