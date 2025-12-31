"""
API для загрузки файлов и парсинга прайс-листов
"""
import io
import os
from typing import List

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Query
from openpyxl import load_workbook

from ..auth import verify_admin_token
from ...services.product_service import ProductService
from ...models.product import ProductCreate
from ...config import settings
from ..utils.excel_parser import ExcelParser
from ...services.image_service import ImageKitService

router = APIRouter()


@router.post("/upload/pricelist")
async def upload_pricelist(
    file: UploadFile = File(...),
    sheet_name: str | None = Form(
        None, description="Название листа для парсинга (один лист)"
    ),
    sheet_index: int = Form(0, description="Индекс листа (0 - первый)"),
    sheet_names: List[str] | None = Form(
        None, description="Список названий листов для парсинга"
    ),
    skip_duplicates: bool = Form(True, description="Пропускать дубликаты"),
    admin: str = Depends(verify_admin_token),
):
    """
    Загрузить и обработать прайс-лист Excel
    
    - Извлекает изображения из ячеек
    - Парсит данные товаров
    - Создает товары в базе данных
    """
    
    if not file.filename:
        raise HTTPException(status_code=400, detail="Файл не выбран")
    
    # Проверяем тип файла
    if not file.filename.lower().endswith('.xlsx'):
        raise HTTPException(
            status_code=400, 
            detail="Поддерживаются только файлы Excel (.xlsx)"
        )
    
    try:
        # Читаем файл
        content = await file.read()
        
        if len(content) > settings.MAX_FILE_SIZE:
            raise HTTPException(
                status_code=400,
                detail=f"Файл слишком большой (макс. {settings.MAX_FILE_SIZE // (1024*1024)}МБ)"
            )
        
        # Очищаем старые временные файлы
        temp_dir = os.path.join(settings.UPLOAD_DIR, "temp_processed_images")
        cleaned_files = ImageKitService.cleanup_temp_folder(temp_dir, max_age_hours=1)
        if cleaned_files > 0:
            print(f"Очищено {cleaned_files} старых временных файлов")

        # Создаем парсер
        parser = ExcelParser(content, settings.UPLOAD_DIR)

        # Парсим файл (один или несколько листов)
        parse_result = parser.parse(
            sheet_name=sheet_name,
            sheet_index=sheet_index,
            sheet_names=sheet_names,
            skip_duplicates=skip_duplicates,
        )
        
        if not parse_result["success"]:
            raise HTTPException(
                status_code=400,
                detail=parse_result.get("error", "Ошибка парсинга")
            )
        
        products_data = parse_result["products"]
        stats = parse_result["stats"]
        errors = parse_result["errors"]

        print(f"[DEBUG] Результат парсинга:")
        print(f"[DEBUG] - Количество товаров: {len(products_data)}")
        print(f"[DEBUG] - Статистика: {stats}")
        print(f"[DEBUG] - Ошибки: {errors}")
        
        # Сохраняем товары в базу
        product_service = ProductService()
        created_count = 0
        save_errors = []
        
        for product_data in products_data:
            try:
                # Создаем товар (image_url уже установлен парсером если найдено изображение)
                product_create = ProductCreate(**product_data)
                created_product = await product_service.create_product(product_create)

                created_count += 1

            except Exception as e:
                error_msg = f"Ошибка сохранения '{product_data.get('name', 'Unknown')}': {str(e)}"
                save_errors.append(error_msg)
        
        return {
            "success": True,
            "message": f"Прайс-лист обработан успешно",
            "sheet_name": parse_result.get("sheet_name"),
            "sheet_names": parse_result.get("sheet_names"),
            "stats": {
                **stats,
                "created": created_count,
                "save_errors": len(save_errors)
            },
            "errors": errors + save_errors[:10]  # Показываем до 10 ошибок
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500, 
            detail=f"Ошибка обработки файла: {str(e)}"
        )


@router.post("/upload/pricelist/sheets")
async def get_pricelist_sheets(
    file: UploadFile = File(...),
    admin: str = Depends(verify_admin_token),
):
    """
    Вернуть список листов Excel-файла прайс-листа.
    Используется фронтендом перед импортом для выбора нужных листов.
    """

    if not file.filename:
        raise HTTPException(status_code=400, detail="Файл не выбран")

    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=400,
            detail="Поддерживаются только файлы Excel (.xlsx)",
        )

    try:
        content = await file.read()

        if len(content) > settings.MAX_FILE_SIZE:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"Файл слишком большой "
                    f"(макс. {settings.MAX_FILE_SIZE // (1024*1024)}МБ)"
                ),
            )

        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=False)
        return {"sheets": list(workbook.sheetnames)}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Ошибка чтения файла Excel: {str(e)}",
        )


@router.get("/upload/template")
async def download_template(admin: str = Depends(verify_admin_token)):
    """Скачать шаблон прайс-листа Excel"""
    
    from fastapi.responses import FileResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    # Создаем новую книгу
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Товары"
    
    # Заголовки
    headers = [
        "Название товара",
        "Фото",
        "Описание/Характеристики",
        "Цена (сум)",
        "Цена ($)",
        "Курс ($)"
    ]
    
    # Стилизуем заголовки
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(1, col_num, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Примеры данных
    example_data = [
        ["IP камера Hikvision DS-2CD1043G0-I 4MP", "", "4Мп IP камера, ИК подсветка до 30м, объектив 2.8мм", 850000, 75, 11333],
        ["Аналоговая камера Dahua HAC-HFW1200TP 2MP", "", "2Мп HDCVI камера, ИК подсветка до 40м", 380000, 33, 11515],
        ["DVR Hikvision DS-7104HGHI-F1 4-канальный", "", "4-канальный видеорегистратор, 1080p", 1200000, 105, 11429]
    ]
    
    for row_num, row_data in enumerate(example_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row_num, col_num, value)
    
    # Настройка ширины столбцов
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    
    # Инструкции на отдельном листе
    ws_instructions = wb.create_sheet("Инструкция")
    
    instructions = [
        ["Инструкция по заполнению прайс-листа"],
        [""],
        ["1. Столбец 'Название товара' - обязательное поле"],
        ["2. Столбец 'Фото' - вставьте изображение в ячейку (Вставка -> Рисунок)"],
        ["3. Столбец 'Описание' - характеристики товара"],
        ["4. Столбец 'Цена (сум)' - цена в узбекских сумах"],
        ["5. Столбец 'Цена ($)' - цена в долларах (опционально)"],
        ["6. Столбец 'Курс ($)' - текущий курс доллара"],
        [""],
        ["Категории определяются автоматически:"],
        ["• IP камеры - содержат: 'IP', 'сетевая', 'PoE'"],
        ["• Аналоговые - содержат: 'AHD', 'HDCVI', 'аналог'"],
        ["• Регистраторы - содержат: 'DVR', 'NVR', 'регистратор'"],
        ["• Аксессуары - содержат: 'блок', 'кабель', 'крепление'"],
        [""],
        ["Советы:"],
        ["• Изображения встраиваются прямо в ячейки Excel"],
        ["• Пустые строки будут пропущены"],
        ["• Дубликаты (одинаковые названия) пропускаются"],
        ["• Цены автоматически пересчитываются если указана одна валюта"]
    ]
    
    for row_num, instruction in enumerate(instructions, 1):
        ws_instructions.cell(row_num, 1, instruction[0])
        if row_num == 1:
            ws_instructions.cell(row_num, 1).font = Font(bold=True, size=14)
    
    ws_instructions.column_dimensions['A'].width = 80
    
    # Сохраняем файл
    temp_file = os.path.join(settings.UPLOAD_DIR, "template_pricelist.xlsx")
    wb.save(temp_file)
    
    return FileResponse(
        temp_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="template_pricelist.xlsx"
    )

